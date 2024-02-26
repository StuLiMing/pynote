#coding:utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
from lxml import etree
import requests
import time,re

def get_html(url):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Cookie': 'SINAGLOBAL=7028353679196.662.1658831650275; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WWcWQDIAu-yR49M31dJ2hIG5JpX5KMhUgL.FoMNSKMpeKq41K52dJLoIp7LxKML1KBLBKnLxKqL1hnLBoMNS0-NeK2c1K.7; PC_TOKEN=c6acc0e342; ALF=1690550615; SSOLoginState=1659014617; SCF=AkOF3RE0iGpbCzsCTVNUUFdnp09eh4AA-S721wBzE0Z9hN2CStPGo6i_A6Y0T8uH9JzJ-sBjIGx4aMRcWtY9vDo.; SUB=_2A25P5v2KDeRhGeFJ7lUQ8SjFwjyIHXVskmhCrDV8PUNbmtAfLW2jkW9Nf4Y14SnjWHSXMXdg3H_cwrf4VO2IdHqK; XSRF-TOKEN=6QGZlTTE4yvDsyfxkk8xRP2B; _s_tentry=weibo.com; Apache=4513244527389.497.1659014654475; ULV=1659014654480:3:3:3:4513244527389.497.1659014654475:1658892594855; WBPSESS=A6VC9miiSVC606I8QigoekS77o0aY48u64XFOar9EGTBgMDJ4ApHpyrcDQzPSEXvT-RV7q5551HqTGbl2UujTookPc5r07RrcMSkfCkQlNX2aa5KyswNPIzcIO7uSuQUQbbA52NnLEy0I98d_IPMPA==',
        'Host': 's.weibo.com',
        'Pragma': 'no-cache',
        'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="90", "Google Chrome";v="90"',
        'sec-ch-ua-mobile': '?0',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'none',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.85 Safari/537.36'
    }
    res = requests.get(url,headers=headers)
    if res.status_code == 200:
        print('获取成功')
        return res.text
    else:
        print('失败')

def jianxi(res):
    data = []
    res = re.findall('<!--card-wrap-->(.*?)<!--/card-wrap-->',res,re.S)
    for r in res:
        xp = etree.HTML(r)
        n = xp.xpath('//p[@class="txt" and @node-type="feed_list_content_full"]//text()')
        if len(n) == 0:
            n = xp.xpath('//p[@class="txt" and @node-type="feed_list_content"]//text()')
        t = xp.xpath('//div[@class="content"]/p[@class="from"]/a[1]/text()')
        p = xp.xpath('//div[@class="card-act"]//li[3]/a/text()')[0]
        d = xp.xpath('//div[@class="card-act"]//li[4]//em/text()')
        if len(d) != 0:
            d = d[0]
        else :
            d = '0'
        p =re.findall('\d*',p)
        p = ''.join('%s' % r.split() for r in p).replace('[', '').replace(']', '').replace('\'', '')
        if p == '':
            p = '0'
        t = ''.join(t[0].split())
        n = ''.join('%s' %r.split() for r in n).replace('[','').replace(']','').replace('\'','')
        n = re.sub(r'\\u...','',n)
        n = re.sub(r'收起全文d','',n)
        data.append({'时间': t, '评论数': p, '点赞数': d, '内容': n})
    return data

def write_data(datas):
    wb = load_workbook('二舅为什么火了.xlsx')
    ws = wb.create_sheet('二舅为什么火了', 0)
    ys = {
        'A':'时间',
        'B':'评论数',
        'C':'点赞数',
        'D':'内容'
    }
    for key, value in ys.items():
        ws[key + '1'] = value
    b = 0
    for data in datas:
        for n in range(len(list(data.values())[0])):
            for key, value in ys.items():
                ws[key + str(n + 2 + b)] = list(data.values())[0][n][value]
        b += len(list(data.values())[0])
    wb.save('二舅为什么火了.xlsx')

if __name__ == '__main__':
    wb = Workbook()
    wb.save('二舅为什么火了.xlsx')
    datas = []
    for i in range(1,51):
        url = 'https://s.weibo.com/weibo?q=%E4%BA%8C%E8%88%85%E4%B8%BA%E4%BB%80%E4%B9%88%E7%81%AB%E4%BA%86'+str(i)
        res = get_html(url)
        data = jianxi(res)
        print(i,data)
        datas.append({str(i): data})
        time.sleep(0.5)
    write_data(datas)
print("over!!!")
